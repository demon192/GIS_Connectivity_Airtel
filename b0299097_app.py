from functools import wraps
from flask import Flask, flash, jsonify, render_template, request, redirect, url_for, session, current_app, Response
import pandas as pd
import numpy as np
import os
import openpyxl
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Change this to a secure secret key

# Define the list of roles
ROLES = ['SuperAdmin', 'DeploymentAdmin', 'GISAdmin', 'DeploymentUser', 'GISUser']

# Read user credentials from the 'demo.xlsx' file
user_credentials = pd.read_excel('user_credentials.xlsx')

file_path = 'demo.xlsx'

excel_file = 'demo.xlsx'  # Change this to the path of your Excel file

df_columns = pd.read_excel('demo.xlsx').columns.tolist()


# def read_excel_data():
#     df = pd.read_excel(excel_file)
#     df.replace({np.nan: ''}, inplace=True)    

#     if(session.get('role') == 'SuperAdmin'):
#         df['Actions'] = df.index.map(lambda idx: f'<div class="btn-group"><button onclick="editRow({idx})" class="edit-btn" >Edit</button><button onclick="deleteRow({idx})" class="delete-btn">Delete</button></div>')
#     elif (session.get('role') in ['DeploymentAdmin','GISAdmin','GISUser']):
#         df['Actions'] = df.apply(lambda row: f'<div class="btn-group"><button onclick="editRow({row.name})" class="edit-btn" {"disabled" if should_disable_submit(row["Final Status"], session.get("role")) else ""}>Edit</button><button onclick="submitChanges({row.name})" class="submit-btn" {"disabled" if should_disable_submit(row["Final Status"], session.get("role")) else ""}>Submit</button></div>', axis=1)
#     else :
#         df['Actions'] = df.index.map(lambda idx: f'<div class="btn-group"><button onclick="editRow({idx})" class="edit-btn" disabled>Edit</button><button onclick="submitChanges({idx})" class="submit-btn" disabled>Submit</button></div>')   

#     # Create a new row for dropdown filters
#     dropdown_row = {}
#     for col in df.columns:
#         dropdown_row[col] = ['All'] + df[col].unique().tolist()  # Include 'All' as default value
#     dropdown_df = pd.DataFrame([dropdown_row])

#     # Concatenate the dropdown row with the existing DataFrame
#     df = pd.concat([dropdown_df, df], ignore_index=True)

#     return df


def read_excel_data():
    df = pd.read_excel(excel_file)
    df.replace({np.nan: ''}, inplace=True)    

    if(session.get('role') == 'SuperAdmin'):
        df['Actions'] = df.index.map(lambda idx: f'<div class="btn-group"><button onclick="editRow({idx})" class="edit-btn" >Edit</button><button onclick="deleteRow({idx})" class="delete-btn">Delete</button></div>')
    elif (session.get('role') in ['DeploymentAdmin','GISAdmin','GISUser']):
        df['Actions'] = df.apply(lambda row: f'<div class="btn-group"><button onclick="editRow({row.name})" class="edit-btn" {"disabled" if should_disable_submit(row["Final Status"], session.get("role")) else ""}>Edit</button><button onclick="submitChanges({row.name})" class="submit-btn" {"disabled" if should_disable_submit(row["Final Status"], session.get("role")) else ""}>Submit</button></div>', axis=1)
    else:
        df['Actions'] = df.index.map(lambda idx: f'<div class="btn-group"><button onclick="editRow({idx})" class="edit-btn" disabled>Edit</button><button onclick="submitChanges({idx})" class="submit-btn" disabled>Submit</button></div>')   

    return df

# Define a list of routes that require authentication
authenticated_routes = ['/index', '/logout', '/add_row', '/submit_row', '/delete_row', '/edit_row', '/download_report', '/get_next_sno', '/add_bulk']  # Add more routes as needed

def should_disable_submit(final_status,role):
    if (role == 'SuperAdmin'):
        return False
    elif (role == 'GISAdmin' or role == 'GISUser'):
        return final_status != 'WIP' 
    elif (role == 'DeploymentAdmin'):
        return final_status != 'Pending'

@app.before_request
def require_login():
    if request.path in authenticated_routes and 'olmId' not in session:
        # Redirect to the login page if the user is not authenticated
        return redirect(url_for('login'))

@app.route('/')
def login():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def process_login():
    olmId = request.form.get('olmId')
    password = request.form.get('password')
    role = request.form.get('role')

    # Convert password to string to ensure consistent comparison
    password = str(password)

    # Remove spaces from the roles read from the Excel file and convert to lowercase
    print('User form ', olmId,password,role)
    print('get :',get_user_detail(olmId,'Password'))
    print('get role: ',get_user_detail(olmId,'Role'))

    print(type(olmId),type(password),type(role))
    print(type(get_user_detail(olmId,'Password')),type(get_user_detail(olmId,'Role')))

    # Check if the provided credentials match any entry in the user credentials DataFrame
    olmId_match = (user_credentials['OLM_ID'] == olmId).any()
    user_password = str(get_user_detail(olmId,'Password'))
    user_role = str(get_user_detail(olmId,'Role')).replace(" ","")

    password_match = (user_password == password)
    role_match = (user_role == role)
    
    # Check if the provided credentials match any entry in the user credentials DataFrame
    if olmId_match and password_match and role_match:
        # Store the credentials in the session
        session['olmId'] = olmId
        session['password'] = password
        session['role'] = role
        return redirect(url_for('index'))
    else:
        flash("Invalid credentials. Please try again.")
        return redirect(url_for('login'))

def role_required(allowed_roles):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Retrieve the role from the session
            role = session.get('role')
            if role in allowed_roles:
                return func(*args, **kwargs)
            else:
                flash("You don't have permission to access this page.", "error")
                return redirect(url_for('index'))  # Redirect to index page or any other page
        return wrapper
    return decorator

# column in demo.xlsx sheet2 is named "Name"
def get_user_detail(olmId, parameter):
    # Read the user credentials from demo.xlsx sheet2
    user_credentials = pd.read_excel('user_credentials.xlsx')
    # Filter the row where OLM_ID matches
    user_row = user_credentials[user_credentials['OLM_ID'] == olmId]
    # Check if a match is found
    if not user_row.empty:
        # Retrieve the name associated with the OLM_ID
        value = user_row.iloc[0][parameter]
        return value
    else:
        return None

# Assuming the name column in demo.xlsx sheet2 is named "Name"
def get_user_name(olmId):
    # Read the user credentials from demo.xlsx sheet2
    user_credentials = pd.read_excel('user_credentials.xlsx')
    # Filter the row where OLM_ID matches
    user_row = user_credentials[user_credentials['OLM_ID'] == olmId]
    # Check if a match is found
    if not user_row.empty:
        # Retrieve the name associated with the OLM_ID
        name = user_row.iloc[0]['Name']
        return name
    else:
        return None

@app.route('/index')
def index():
    # Retrieve the role from the session
    role = session.get('role')
    olmId = session.get('olmId')
    password = session.get('password')

    if olmId and password and role:
        # Read the Excel file from the server's filesystem
        df = read_excel_data()
        # Retrieve the user's name
        name = get_user_name(olmId)
        # Render the HTML template with the Excel data and the user's role
        return render_template('index.html',
                               html_table=df.to_html(index=False,escape=False),name = name, olmId=olmId, password=password, role=role)
    else:
        return redirect(url_for('login'))
    

@app.route('/pending_tasks')
def pending_tasks():
    # Retrieve the role, olmId, and password from the session
    role = session.get('role')
    olmId = session.get('olmId')
    password = session.get('password')

    if olmId and password and role:
        # Read the Excel file from the server's filesystem
        df = read_excel_data()
        # Retrieve the user's name
        name = get_user_name(olmId)
        name_lower = name.upper()
        # Filter the DataFrame to show only rows where 'Responsibility' matches the user's name
        df['Responsibility'] = df['Responsibility'].str.upper()
        filtered_df = df[df['Responsibility'] == name_lower]
        filtered_df = filtered_df[filtered_df['Final Status'] == 'WIP']
        # Render the HTML template with the filtered DataFrame and the user's role
        return render_template('index.html',
                               html_table=filtered_df.to_html(index=False, escape=False), name=name, olmId=olmId, password=password, role=role)
    else:
        return redirect(url_for('login'))

@app.route('/logout')
def logout():
    # Clear the session to log out the user
    session.pop('olmId', None)
    session.pop('password', None)
    session.pop('role', None)

    return redirect(url_for('login'))

@app.route('/delete_row', methods=['POST'])
def delete_row():
    # Step 1: Check the user's role
    role = session.get('role')
    print("User's role:", role)  # Debug print statement
    if role not in ['SuperAdmin']:
        # Redirect or display an error message
        flash("You don't have permission to Delete Row", "error")
        return redirect(url_for('index'))
    
    index = request.json.get('index')
    print('index: ',index)
    if index is not None:
        try:
            # Load the Excel file and delete the row
            df = pd.read_excel(excel_file)
            df.drop(index=index, inplace=True)
            df.to_excel(excel_file, index=False)  # Save the updated DataFrame to the Excel file
            return '', 204  # HTTP status code for successful deletion
        except Exception as e:
            print('Error deleting row:', e)
            return 'Error deleting row', 500  # HTTP status code for server error
    else:
        return 'Index parameter missing', 400  # HTTP status code for bad request

@app.route('/edit_row/<int:index>')
def edit_row(index):
    # Step 1: Check the user's role
    role = session.get('role')
    df = read_excel_data()  # Read the Excel file

    # print("User's role:", role)  # Debug print statement
    final_status = df.iloc[index]['Final Status']
    if role not in ['SuperAdmin', 'GISAdmin', 'GISUser']:
        # Redirect or display an error message
        if role == 'DeploymentAdmin' and (final_status == 'WIP' or final_status == 'Done'):
            flash("You don't have permission to Edit Row.", "error")
            return redirect(url_for('index'))

    row_data = df.iloc[index].to_dict()  # Get the data of the row to be edited
    actions_hidden = True  # Flag indicating whether the 'Actions' column should be hidden
    return render_template('edit_row.html', row_data=row_data, index=index, actions_hidden=actions_hidden, role=role)

@app.route('/submit_edit/<int:index>', methods=['POST'])
def submit_edit(index):
    olmId = session.get('olmId')
    role = get_user_detail(olmId,'Role')
    row_data = {column: request.form[column] for column in df_columns}
    df = read_excel_data()  # Read the Excel file
    df.iloc[index] = row_data  # Update the row with the edited values

    mandatory_gis_rows = ['A End M6 Code', 'Site-A Mux ID', 'Fiber distance from A-end (Km)', 'B End M6 Code', 'Site-B Mux ID', 'Fiber distance from Z-end (Km)', 'Fiber Owner', 'Ring Name(If available)', 'NO. Of node available in Ring', 'Ring utilization in %', 'GIS REMARKS', 'Remarks']

    mandatory_deployment_rows = ['Dep. S.No.', 'Partner', 'HUB', 'Circle', 'Parent Site', 'Parent Site UID', 'Alternate', 'Category', 'Final Site', 'M6 Code', 'LAT', 'LONG']
    # Update the 'Status' column value to 'WIP' for the edited row
    print('role1 :',role)
    if role in ['GIS Admin', 'GIS User', 'SuperAdmin']:
        if all(value for key, value in row_data.items() if key in mandatory_gis_rows):
            # Update the 'Status' column value to 'Done' if all fields are filled except the two specified
            df.at[index, 'Status'] = 'Done'
        elif df.at[index, 'Status'] in ['Not Feasible', 'Route Not Available in ICW', 'Site Already Fiberized'] : 
                print('role: ',role)
                print('Status: ',df.at[index,'Status'])
                df.at[index, 'Final Status'] = 'Done'
        elif df.at[index, 'Status'] == 'Intracity':
                df.at[index, 'Final Status'] = 'Intracity'
        else: 
            df.at[index,'Final Status'] = 'WIP'

    if (role == 'DeploymentAdmin' or role == 'SuperAdmin') and df.at[index, 'Status'] == '' :
        if all(value for key, value in row_data.items() if key in mandatory_deployment_rows):
            df.at[index, 'Final Status'] = 'WIP'
        else:
            df.at[index, 'Final Status'] = 'Pending'
    # Update the validation date to current time stamp
    df.at[index,'Validation Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # Update the Responsibility to Gis User or GIS Admin
    if(role != 'SuperAdmin'):
        df.at[index,'Responsibility'] = str(get_user_detail(olmId,'Name'))
    df.to_excel(excel_file, index=False)  # Save the updated DataFrame to the Excel file
    return redirect(url_for('index'))

@app.route('/submit_changes/<int:index>', methods=['POST'])
def submit_changes(index):
    role = session.get('role')
    df = read_excel_data()
    row_data = df.iloc[index].to_dict()
    final_status = row_data['Final Status']
    index = request.json.get('index')

    mandatory_gis_rows = ['A End M6 Code', 'Site-A Mux ID', 'Fiber distance from A-end (Km)', 'B End M6 Code', 'Site-B Mux ID', 'Fiber distance from Z-end (Km)', 'Fiber Owner', 'Ring Name(If available)', 'NO. Of node available in Ring', 'Ring utilization in %', 'Status', 'Remarks']

    mandatory_deployment_rows = ['Proirity', 'HUB', 'Circle', 'Final Site', 'M6 Code', 'LAT', 'LONG', 'Address', 'SLC', 'Planning Town Category', 'Node Type', 'Route Name']

    if index is not None:
        try:
            if ( role in ['GISAdmin', 'GISUser']):
                # Update the 'Status' column value to 'Done' if all fields are filled except the two specified
                if all(value for key, value in row_data.items() if key in mandatory_gis_rows):
                    df.at[index, 'Final Status'] = 'Done'
                    flash("Submission Done Successfully", "success")
                else:
                    flash("Please fill all the required fields first", "error")

            elif (role == 'DeploymentAdmin'):
                
                if all(value for key, value in row_data.items() if key in mandatory_deployment_rows):
                # Otherwise, update the 'Status' column value to 'WIP'
                    flash("Submission Done Successfully","success")
                    df.at[index, 'Final Status'] = 'WIP'
                else:
                    flash("Please fill all the required fields first","error")
            else:
                    df.at[index, 'Final Status'] = 'Pending'
                    flash("No changes","error")
            df.to_excel(excel_file, index=False)  # Save the updated DataFrame to the Excel file
            # return redirect(url_for('index'))
            return '', 204  # HTTP status code for successful deletion
        except Exception as e:
            print('Error Submitting row:', e)
            return 'Error Submitting row', 500  # HTTP status code for server error
    else:
        return 'Index parameter missing', 400  # HTTP status code for bad request 

@app.route('/add_row')
# @role_required(['SuperAdmin', 'DeploymentAdmin'])
def add_row():
    # Step 1: Check the user's role
    role = session.get('role')
    print("User's role:", role)  # Debug print statement
    if role not in ['SuperAdmin', 'DeploymentAdmin']:
        # Redirect or display an error message
        flash("You don't have permission to ADD Rows.", "error")
        return redirect(url_for('index'))
    
    return render_template('add_row.html', columns=df_columns,hide_actions=True)

@app.route('/submit_row', methods=['POST'])
def submit_row():
    # Retrieve the role from the session
    role = session.get('role')
    olmId = session.get('olmId')
    if role:
        row_data = {column: request.form[column] for column in df_columns if column != 'Actions'}

        # Set status to "Pending"
        row_data['Final Status'] = 'Pending'

        new_row = pd.DataFrame(row_data, index=[0])

        if os.path.exists(excel_file):
            # If the file already exists, load the existing workbook
            wb = openpyxl.load_workbook(excel_file)
            if 'Sheet1' in wb.sheetnames:
                # If 'Sheet1' exists, get its worksheet
                ws = wb['Sheet1']
                # Append the new row
                for index, row in new_row.iterrows():
                    # Convert the current date to a string in Excel format before adding to the DataFrame
                    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    row['Received Date'] = current_date 
                    row['Received From'] = str(get_user_detail(olmId,'Name'))
                    ws.append(row.tolist())
            else:
                # If 'Sheet1' doesn't exist, create a new workbook and add 'Sheet1'
                ws = wb.create_sheet('Sheet1')
                # Write the header row
                for col, column in enumerate(df_columns, start=1):
                    ws.cell(row=1, column=col, value=column)
                # Write the new row
                for index, row in new_row.iterrows():
                    ws.append(row.tolist())
            # Save the changes
            wb.save(excel_file)
        else:
            # If the file doesn't exist, create a new file with the new row
            new_row.to_excel(excel_file, index=False, header=True)

        return redirect(url_for('index'))
    else:
        return redirect(url_for('login'))

@app.route('/get_next_sno')
def get_next_sno():
    try:
        df = pd.read_excel(excel_file)
        if not df.empty:
            last_sno = df['Unique Ref'].iloc[-1]
            next_sno = int(last_sno.split('GIS')[-1]) + 1
        else:
            next_sno = 1
        next_sno_str = f"GIS{next_sno:06d}"  # Format the next serial number
        return jsonify({'next_sno': next_sno_str})
    except Exception as e:
        print('Error getting next serial number:', e)
        return jsonify({'next_sno': 'GIS000001'})  # Default value if an error occurs

@app.route('/download_report')
def download_report():
    try:
        # Check if demo.xlsx exists in the current directory
        if os.path.exists('demo.xlsx'):
            # Generate the filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"GIS_DEPLOYMENT_REPORT_{timestamp}.xlsx"

            # Read the Excel file without the 'Actions' column
            df = pd.read_excel('demo.xlsx')
            df = df.drop(columns=['Actions'])

            # Save the DataFrame to a new Excel file without the 'Actions' column
            df.to_excel('temp.xlsx', index=False)

            # Return the new Excel file as a downloadable attachment with the custom filename
            with open('temp.xlsx', 'rb') as file:
                response = Response(file.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response.headers['Content-Disposition'] = f"attachment; filename={filename}"
                return response
        else:
            return "Error: demo.xlsx file not found in the current directory", 404
    except Exception as e:
        current_app.logger.error(f"Error downloading report: {e}")
        return "Error downloading report: please check the server logs for details", 500
    
@app.route('/download_pending_report')
def download_pending_report():
    try:
        # Check if demo.xlsx exists in the current directory
        if os.path.exists('demo.xlsx'):
            olmId = session.get('olmId')
            name = get_user_name(olmId)
            # Generate the filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{name}_PENDING_REPORT_{timestamp}.xlsx"
            # Read the Excel file without the 'Actions' column
            df = pd.read_excel('demo.xlsx')

            name_lower = name.upper()

            df = df.drop(columns=['Actions'])
            df['Responsibility'] = df['Responsibility'].str.upper()
            filtered_df = df[df['Responsibility'] == name_lower]
            filtered_df = filtered_df[filtered_df['Final Status'] == 'WIP']

            # Save the DataFrame to a new Excel file without the 'Actions' column
            filtered_df.to_excel('temp.xlsx', index=False)

            # Return the new Excel file as a downloadable attachment with the custom filename
            with open('temp.xlsx', 'rb') as file:
                response = Response(file.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response.headers['Content-Disposition'] = f"attachment; filename={filename}"
                return response
        else:
            return "Error: demo.xlsx file not found in the current directory", 404
    except Exception as e:
        current_app.logger.error(f"Error downloading report: {e}")
        return "Error downloading report: please check the server logs for details", 500


@app.route('/add_bulk', methods=['GET','POST'])
def add_bulk():
    # Step 1: Check the user's role
    role = session.get('role')
    olmId = session.get('olmId')
    print("User's role:", role)  # Debug print statement
    if role not in ['SuperAdmin', 'DeploymentAdmin']:
        # Redirect or display an error message
        flash("You don't have permission to ADD Bulk.", "error")
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Get the uploaded file
        uploaded_file = request.files['file']
        if uploaded_file.filename == '':
            return "No file selected", 400
        try:
            # Read the uploaded file into a DataFrame
            df_uploaded = pd.read_excel(uploaded_file)

            # Exclude "Actions" column from comparison
            df_demo = pd.read_excel('demo.xlsx')
            df_demo_columns = [col for col in df_demo.columns if col != "Actions"]

            # Check if column names (excluding "Actions") match
            if set(df_demo_columns) != set(df_uploaded.columns):
                return "Column names do not match", 400
            
            print('1')
            # Remove the existing 'Unique Ref' column if it exists
            if 'Unique Ref' in df_uploaded.columns:
                df_uploaded.drop(columns=['Unique Ref'], inplace=True)
            print('2')
            # Generate Unique Ref values for the new rows
            start_index = len(df_demo) + 1
            end_index = start_index + len(df_uploaded)
            s_no_values = ['GIS{:06d}'.format(i) for i in range(start_index, end_index)]
            print('3')
            # Insert the Unique Ref column to the DataFrame
            df_uploaded.insert(0, 'Unique Ref', s_no_values)
            print('4')
            # Load the existing workbook
            wb = load_workbook('demo.xlsx')
            print('5')
            ws = wb.active
            # Append the data to the worksheet
            for row in df_uploaded.values:
                ws.append(row.tolist())

            # Select the active worksheet
            if( role != 'SuperAdmin'):
                # Clear values of columns beyond the 15th column for all appended rows
                for row in ws.iter_rows(min_row=ws.max_row - len(df_uploaded) + 1, max_row=ws.max_row, min_col=16):
                    for cell in row:
                        cell.value = None

                # Find the index of 'Status' column
                final_status_column_index = df_demo.columns.get_loc('Final Status') if 'Final Status' in df_demo.columns else -1

                # Set the value of 'Status' column to 'Pending' for all appended rows
                if final_status_column_index != -1:
                    for cell in ws.iter_rows(min_row=ws.max_row - len(df_uploaded) + 1, max_row=ws.max_row, min_col=final_status_column_index + 1, max_col=final_status_column_index + 1):
                        cell[0].value = 'Pending'

            # Find the index of 'Recived Date' column
            received_date_column_index = df_demo.columns.get_loc('Received Date') if 'Received Date' in df_demo.columns else -1
            
            # Set the value of 'Recived Date' column to 'Current Time Stamp' for all appended rows
            if received_date_column_index != -1:
                for row in ws.iter_rows(min_row=ws.max_row - len(df_uploaded) + 1, max_row=ws.max_row, min_col=received_date_column_index + 1, max_col=received_date_column_index + 1):
                    for cell in row:
                        cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Find the index of 'Recived From' column
            received_from_column_index = df_demo.columns.get_loc('Received From') if 'Received From' in df_demo.columns else -1

            # Set the value of 'Received From' column to 'Pending' for all appended rows
            if received_from_column_index != -1:
                for row in ws.iter_rows(min_row=ws.max_row - len(df_uploaded) + 1, max_row=ws.max_row, min_col=received_from_column_index + 1, max_col=received_from_column_index + 1):
                    for cell in row:
                        cell.value = str(get_user_detail(olmId,'Name'))

            # Save the changes to the Excel file
            wb.save('demo.xlsx')

            # Redirect back to the index page
            return redirect(url_for('index'))
        except Exception as e:
            print("Error:", e)
            return "An error occurred", 500
    else:
        return render_template('add_bulk.html')

if __name__ == '__main__':
    app.run(debug=True)